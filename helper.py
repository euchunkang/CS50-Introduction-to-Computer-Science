import requests
from openpyxl import load_workbook
from flask import render_template

api_key = 'AIzaSyAlAh18il2GWEWw7L20ORFzoiEXzKOcSV4'

# from excel sheet, get list of addresses
def get_address(excel_file):
    wb = load_workbook(excel_file)
    ws = wb['add']
    # extract addresses from Excel Sheet
    add_list = []
    for row in ws.iter_cols():
        if row[0].value == 'Addresses':
            j = 0
            for data in row[1:]:
                add_list.append(str(data.value))
            break
    return add_list

# from address list, get coordinates
def get_coordinates(address_list):
    latitudes = []
    longitudes = []
    formatted_address = []
    for address in address_list:
        r = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address="%s"&key=%s'
                         %(address,api_key))
        result = r.json()["results"]
        location = result[0]["geometry"]["location"]
        latitudes.append(location["lat"])
        longitudes.append(location["lng"])
        formatted_address.append(result[0]["formatted_address"])
    return latitudes, longitudes, formatted_address

# get optimised route
def best_route(startpoint, endpoint, address_list):
    address_string = '|'.join(address_list)
    r = requests.get(
        'https://maps.googleapis.com/maps/api/directions/json?origin=%s&destination=%s&waypoints=optimize:true|%s&key=%s'
        %(startpoint, endpoint, address_string, api_key))
    info = r.json()
    order = info['routes'][-1]['waypoint_order']
    return order

#failure page
def failure(error_msg) -> object:
    return render_template("failure.html", error_msg=error_msg)

def get_url(address_list):
    final_url = "https://www.google.com/maps/dir/"
    for i in range(len(address_list)):
        x = address_list[i]
        final_url += x.replace(" ","+") + "/"
    return final_url

